# -*- coding: utf-8 -*-

from odoo import http
from odoo.http import request
from odoo.addons.web.controllers.main import content_disposition
import base64
from io import BytesIO
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except ImportError:
	install('openpyxl==3.1.5')

class Download_xls(http.Controller):

	@http.route('/web/binary/download_template_update_gratification/<int:gratification_id>', type='http', auth="public")
	def download_template_update_gratification(self, gratification_id, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_gratification.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Importador de Gratificaciones.xlsx'
		# filecontent = base64.b64decode(filecontent)
		# workbook = openpyxl.load_workbook(BytesIO(filecontent))

		infile = BytesIO()
		infile.write(base64.b64decode(filecontent))
		infile.seek(0)
		workbook = openpyxl.load_workbook(filename=infile)

		sheet = workbook['IMPORTADOR']
		gratification = request.env['hr.gratification'].sudo().search([('id', '=', gratification_id)])
		row = 2
		if gratification:
			for i in gratification.line_ids:
				sheet[f'A{row}'] = i.identification_id
				sheet[f'B{row}'] = i.id
				sheet[f'C{row}'] = i.employee_id.name
				sheet[f'D{row}'] = i.months
				sheet[f'E{row}'] = i.days
				sheet[f'F{row}'] = i.lacks
				sheet[f'G{row}'] = i.wage
				sheet[f'H{row}'] = i.household_allowance
				sheet[f'I{row}'] = i.commission
				sheet[f'J{row}'] = i.bonus
				sheet[f'K{row}'] = i.extra_hours
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])


	@http.route('/web/binary/download_template_update_cts/<int:cts_id>', type='http', auth="public")
	def download_template_update_cts(self, cts_id, **kw):

		invoice_xls = request.env['ir.attachment'].sudo().search([('name', '=', 'sample_file_update_cts.xlsx')])
		filecontent = invoice_xls.datas
		filename = 'Plantilla Importador de CTS.xlsx'
		# filecontent = base64.b64decode(filecontent)
		# workbook = openpyxl.load_workbook(BytesIO(filecontent))

		infile = BytesIO()
		infile.write(base64.b64decode(filecontent))
		infile.seek(0)
		workbook = openpyxl.load_workbook(filename=infile)

		sheet = workbook['IMPORTADOR']
		cts = request.env['hr.cts'].sudo().search([('id', '=', cts_id)])
		row = 2
		if cts:
			for i in cts.line_ids:
				sheet[f'A{row}'] = i.identification_id
				sheet[f'B{row}'] = i.id
				sheet[f'C{row}'] = i.employee_id.name
				sheet[f'D{row}'] = i.months
				sheet[f'E{row}'] = i.days
				sheet[f'F{row}'] = i.lacks
				sheet[f'G{row}'] = i.wage
				sheet[f'H{row}'] = i.household_allowance
				sheet[f'I{row}'] = i.sixth_of_gratification
				sheet[f'J{row}'] = i.commission
				sheet[f'K{row}'] = i.bonus
				sheet[f'L{row}'] = i.extra_hours
				row += 1

		output_stream = BytesIO()
		workbook.save(output_stream)

		output_content = output_stream.getvalue()

		return request.make_response(output_content,
									 [('Content-Type', 'application/octet-stream'),
									  ('Content-Disposition', content_disposition(filename))])
