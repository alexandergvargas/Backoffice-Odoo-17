# -*- coding: utf-8 -*-
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)
import base64
from io import BytesIO
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')


class LandedCostIt(models.Model):
	_inherit = 'landed.cost.it'

	move_ids = fields.One2many(
		string=_('Asientos/Facturas'),
		comodel_name='account.move',
		inverse_name='landed_cost_id',
		copy=False
	)
	
	move_entry_ids = fields.One2many(
		string=_('Asientos generados'),
		comodel_name='account.move',
		inverse_name='landed_cost_id',
		copy=False,
		domain=[('move_type', '=', 'entry')]
	)
	

	
	def borrador(self):
		for record in self:
			if record.move_entry_ids:
				raise UserError(u'Debe de eliminar todos sus asientos generados para establecer a Borrador el GV')
		return super(LandedCostIt, self).borrador()


	def get_info(self):
		res = super(LandedCostIt, self).get_info()
		for record in self:
			invoices = self.env['account.move'].search([('landed_cost_id','=',self.id),('state','=','posted')])
			for move in invoices:
				for line in move.line_ids:
					if line.product_id.is_landed_cost:
						vals = {
							'invoice_id': line.id,
							'invoice_date': line.move_id.invoice_date,
							'type_document_id': line.type_document_id.id,
							'nro_comp': line.nro_comp,
							'date': line.move_id.date,
							'partner_id': line.partner_id.id,
							'product_id': line.product_id.id,
							'debit': (line.debit - line.credit),
							'amount_currency': (line.debit - line.credit)/line.tc if line.move_id.currency_id.id == line.company_id.currency_id.id else line.amount_currency,
							'tc': line.tc,
							'type_landed_cost_id': line.product_id.type_landed_cost_id.id,
							'company_id': line.company_id.id,
						}
						record.write({'invoice_ids' :([(0,0,vals)]) })
						record._change_flete()		
		return res

	def open_entries(self):
		self.ensure_one()
		action = self.env.ref('account.action_move_journal_line').read()[0]
		domain = [('id', 'in', self.move_entry_ids.ids)]
		context = dict(self.env.context, default_invoice_id=self.id)
		views = [(self.env.ref('account.view_move_tree').id, 'tree'), (False, 'form'), (False, 'kanban')]
		return dict(action, domain=domain, context=context, views=views)
	
	def generate_excel(self):
		import io
		from xlsxwriter.workbook import Workbook
		from xlsxwriter.utility import xl_rowcol_to_cell

		ReportBase = self.env['report.base']
		direccion = self.env['account.main.parameter'].search([('company_id','=',self.company_id.id)],limit=1).dir_create_file

		if not direccion:
			raise UserError(u'No existe un Directorio Exportadores configurado en Parametros Principales de Contabilidad para su Compañía')

		workbook = Workbook(direccion +'Reporte_GV.xlsx')
		workbook, formats = ReportBase.get_formats(workbook)

		numbertotalocho = workbook.add_format({'num_format':'0.00000000','bold': True})
		numbertotalocho.set_align('right')
		numbertotalocho.set_align('vcenter')
		numbertotalocho.set_border(style=1)
		numbertotalocho.set_font_size(10.5)
		numbertotalocho.set_font_name('Times New Roman')
		numbertotalocho.set_underline()

		import importlib
		import sys
		importlib.reload(sys)

		worksheet = workbook.add_worksheet("GV")
		worksheet.set_tab_color('blue')

		HEADERS = ['REFERENCIA','DE','PARA','PRODUCTO','UNIDAD','CANTIDAD','P.UNIT','VALOR','FACTOR','GASTO V','ADVALOREM','VALOR TOTAL','C.UNIT']
		worksheet = ReportBase.get_headers(worksheet,HEADERS,0,0,formats['boldbord'])

		x=1
		init = 1

		for line in self.detalle_ids:
			worksheet.write(x,0,line.picking_rel.display_name if line.picking_rel else '',formats['especial1'])
			worksheet.write(x,1,line.origen_rel.display_name if line.origen_rel else '',formats['especial1'])
			worksheet.write(x,2,line.destino_rel.display_name if line.destino_rel else '',formats['especial1'])
			worksheet.write(x,3,line.producto_rel.display_name if line.producto_rel else '',formats['especial1'])
			worksheet.write(x,4,line.unidad_rel.display_name if line.unidad_rel else '',formats['especial1'])
			worksheet.write(x,5,line.cantidad_rel if line.cantidad_rel else 0,formats['numberdos'])
			worksheet.write(x,6,line.precio_unit_signed if line.precio_unit_signed else 0,formats['numberocho'])
			worksheet.write(x,7,line.valor_rel_signed if line.valor_rel_signed else 0,formats['numberdos'])
			worksheet.write(x,8,line.factor if line.factor else '',formats['numberocho'])
			worksheet.write(x,9,line.flete if line.flete else '',formats['numberocho'])
			worksheet.write(x,10,line.advalorem if line.advalorem else '',formats['numberdos'])
			worksheet.write(x,11,line.total if line.total else '',formats['numberdos'])
			worksheet.write(x,12,line.total/line.cantidad_rel if line.cantidad_rel and line.cantidad_rel != 0 else 0,formats['numberocho'])
			x += 1

		worksheet.write_formula(x,5, '=SUM(' + xl_rowcol_to_cell(init,5) + ':' + xl_rowcol_to_cell(x-1,5) + ')', formats['numbertotal'])
		worksheet.write_formula(x,7, '=SUM(' + xl_rowcol_to_cell(init,7) + ':' + xl_rowcol_to_cell(x-1,7) + ')', formats['numbertotal'])
		worksheet.write_formula(x,9, '=SUM(' + xl_rowcol_to_cell(init,9) + ':' + xl_rowcol_to_cell(x-1,9) + ')', numbertotalocho)
		worksheet.write_formula(x,10, '=SUM(' + xl_rowcol_to_cell(init,10) + ':' + xl_rowcol_to_cell(x-1,10) + ')', formats['numbertotal'])
		worksheet.write_formula(x,11, '=SUM(' + xl_rowcol_to_cell(init,11) + ':' + xl_rowcol_to_cell(x-1,11) + ')', formats['numbertotal'])

		widths = [15,25,20,30,15,15,20,20,20,20,20,20,17]
		worksheet = ReportBase.resize_cells(worksheet,widths)
		workbook.close()

		f = open(direccion +'Reporte_GV.xlsx', 'rb')

		return self.env['popup.it'].get_file('Reporte GV.xlsx',base64.encodebytes(b''.join(f.readlines())))
	
	def get_sql_report_saldos(self):
		sql = sql1 = sqlsum = ""
		sql2 = """valor_p + valormn """
		sql3 = """ valormn """
		for elem in self.env['landed.cost.it.type'].search([]):
			sql2 += "+"
			sql3 += "+"
			sql2 += """ coalesce("%s",0) """%(elem.code)
			sql3 += """ coalesce("%s",0) """%(elem.code)
			sql += ", \n"
			sql1 += """ coalesce("%s",0) as "%s", """%(elem.code, elem.name)
			sqlsum += """ sum(coalesce("%s",0)) as "%s", """%(elem.code, elem.name)
			sql += """ a.factor*(select sum(debit) from landed_cost_invoice_line where landed_id=%d and type_landed_cost_id = %d) as "%s" """%(self.id,elem.id,elem.code)

		sql = """(select almacen,albaran,codigo,producto,cantidad,factor,valor_p, valormn as "Advalorem ", 
		%s
								%s as total_gv,
								%s as costo_total,
								(%s)/valor_p as factor_d,
								(%s)/cantidad as costo_unitario
								from 
								(
								select   
								b.almacen, 
								b.albaran,
								b.default_code as codigo, 
								b.name_template as producto, 
								b.ingreso as cantidad,
								a.factor,
								b.debit as valor_p,
								coalesce(c.valormn,0) as valormn %s
								from landed_cost_it_line a
								LEFT JOIN 
								(
								select sl.complete_name as almacen,
								sp.name as albaran,
								line.stock_move_id as stock_moveid,
								sm.product_id,
								pp.default_code,
								(pt.name->>'es_PE')::character varying as name_template,
								sum(sm.product_qty) as ingreso,
								sum(line.valor_rel_signed) as debit
								from landed_cost_it_line line
								left join stock_move sm on sm.id = line.stock_move_id
								left join stock_picking sp on sp.id = sm.picking_id
								left join stock_location sl on sl.id = sm.location_dest_id
								left join product_product pp on pp.id = sm.product_id
								left join product_template pt on pt.id = pp.product_tmpl_id
								where line.gastos_id = %d
								group by sl.complete_name, sp.name, line.stock_move_id, sm.product_id, pp.default_code, (pt.name->>'es_PE')::character varying
								order by sl.complete_name, sp.name
								) b ON b.stock_moveid = a.stock_move_id
								LEFT JOIN stock_move SM on SM.id = b.stock_moveid
								LEFT JOIN (select landed_id, picking_id , product_id, sum(coalesce(valormn,0)) as valormn from landed_cost_advalorem_line GROUP BY landed_id, picking_id, product_id) c ON c.product_id = b.product_id AND a.gastos_id = c.landed_id and SM.picking_id = c.picking_id
								where a.gastos_id = %d
								)tt)
								UNION ALL
								(select null as almacen,null as albaran,null as codigo,null as producto,null as cantidad, null as factor,sum(tt2.valor_p) as valor_p, sum(tt2.valormn) as "Advalorem ", 
								%s
								sum(%s) as total_gv,
								sum(%s) as costo_total,
								null as factor_d,
								null as costo_unitario
								from 
								(
								select   
								b.debit as valor_p,
								coalesce(c.valormn,0) as valormn %s
								from landed_cost_it_line a
								LEFT JOIN 
								(
									
								select line.stock_move_id as stock_moveid,
								sm.product_id,
								sum(line.valor_rel_signed) as debit
								from landed_cost_it_line line
								left join stock_move sm on sm.id = line.stock_move_id
								left join stock_location sl on sl.id = sm.location_dest_id
								left join product_product pp on pp.id = sm.product_id
								left join product_template pt on pt.id = pp.product_tmpl_id
								where line.gastos_id = %d
								group by line.stock_move_id, sm.product_id
								) b ON b.stock_moveid = a.stock_move_id
								LEFT JOIN stock_move SM on SM.id = b.stock_moveid
								LEFT JOIN (select landed_id , picking_id, product_id, sum(coalesce(valormn,0)) as valormn from landed_cost_advalorem_line GROUP BY landed_id, picking_id, product_id) c ON c.product_id = b.product_id AND a.gastos_id = c.landed_id and SM.picking_id = c.picking_id
								where a.gastos_id = %d
								)tt2)"""%(sql1,
									sql3,
									sql2,
									sql2,
									sql2,
									sql,
									self.id,
									self.id,
									sqlsum,
									sql3,
									sql2,
									sql,
									self.id,
									self.id
								)
		return sql
	 
	def get_excel_saldos(self):
		self.ensure_one()
		self.env.cr.execute(self.get_sql_report_saldos())
		res = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		res.insert(0, colnames)

		wb = openpyxl.Workbook()
		ws = wb.active
		row_position = 1
		col_position = 1
		for index, row in enumerate(res, row_position):
			for col, val in enumerate(row, col_position):
				ws.cell(row=index, column=col).value = val
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()

		return self.env['popup.it'].get_file('%s.xlsx'%(self.name),output_datas)

	def get_sql_saldos(self):
		sql = sql1 = sqlsum = ""
		sql2 = """valor_p + valormn """
		sql3 = """ valormn """
		for elem in self.env['landed.cost.it.type'].search([]):
			sql2 += "+"
			sql3 += "+"
			sql2 += """ coalesce("%s",0) """%(elem.code)
			sql3 += """ coalesce("%s",0) """%(elem.code)
			sql += ", \n"
			sql1 += """ round(coalesce("%s"::numeric,0),2) as "%s", """%(elem.code, elem.name)
			sqlsum += """ sum(coalesce("%s",0)) as "%s", """%(elem.code, elem.name)
			sql += """ a.factor*(select sum(debit) from landed_cost_invoice_line where landed_id=%d and type_landed_cost_id = %d) as "%s" """%(self.id,elem.id,elem.code)

		sql = """(select almacen,albaran,fecha_albaran,stock_moveid,codigo,producto,cantidad,factor, round(valor_p::numeric,2) as valor_p, round(valormn,2) as "Advalorem ", 
		%s
								%s as total_gv,
								%s as costo_total,
								(%s)/valor_p as factor_d,
								(%s)/cantidad as costo_unitario
								from 
								(
								select   
								b.almacen, 
								b.albaran, 
								b.fecha_albaran,
								b.stock_moveid,
								b.default_code as codigo, 
								b.name_template as producto, 
								b.ingreso as cantidad,
								a.factor,
								b.debit as valor_p,
								coalesce(c.valormn,0) as valormn %s
								from landed_cost_it_line a
								LEFT JOIN 
								(
								select sl.complete_name as almacen,
								sp.name as albaran,
								(sp.kardex_date::timestamp - interval '5' hour)::date as fecha_albaran,
								line.stock_move_id as stock_moveid,
								sm.product_id,
								pp.default_code,
								(pt.name->>'es_PE')::character varying as name_template,
								sum(sm.product_qty) as ingreso,
								sum(line.valor_rel_signed) as debit
								from landed_cost_it_line line
								left join stock_move sm on sm.id = line.stock_move_id
								left join stock_picking sp on sp.id = sm.picking_id
								left join stock_location sl on sl.id = sm.location_dest_id
								left join product_product pp on pp.id = sm.product_id
								left join product_template pt on pt.id = pp.product_tmpl_id
								where line.gastos_id = %d
								group by sl.complete_name, sp.name, sp.kardex_date, line.stock_move_id, sm.product_id, pp.default_code, (pt.name->>'es_PE')::character varying
								order by sl.complete_name, sp.name
								) b ON b.stock_moveid = a.stock_move_id
								LEFT JOIN stock_move SM on SM.id = b.stock_moveid
								LEFT JOIN (select landed_id, picking_id , product_id, sum(coalesce(valormn,0)) as valormn from landed_cost_advalorem_line GROUP BY landed_id, picking_id, product_id) c ON c.product_id = b.product_id AND a.gastos_id = c.landed_id and SM.picking_id = c.picking_id
								where a.gastos_id = %d
								order by b.almacen, b.albaran
								)tt)"""%(sql1,
									sql3,
									sql2,
									sql2,
									sql2,
									sql,
									self.id,
									self.id
								)
		return sql

	def action_create_moves(self):
		param = self.env['account.main.parameter'].search([('company_id','=',self.env.company.id)],limit=1)
		if not param.lc_account_id:
			raise UserError(u'No esta configurada la Cuenta de Ingreso de Existencias por recibir en los Parámetros principales de Contabilidad en la Compañía')
		if not param.lc_journal_id:
			raise UserError(u'No esta configurado el Diario de Ingreso de Existencias por recibir en los Parámetros principales de Contabilidad en la Compañía')
		if not param.others_document_type:
			raise UserError(u'No esta configurado el Tipo Documento Otros en los Parámetros principales de Contabilidad en la Compañía')
		for landed in self:
			if landed.state != 'done':
				raise UserError(u'No puede aplicar esta acción si el GV no se encuentra en estado Finalizado.')
			if landed.move_entry_ids:
				raise UserError(u'Ya tiene asientos generados, procure borrarlos todos para volver a generar la acción')
			lines_without_tgv = landed.invoice_ids.filtered(lambda l: not l.type_landed_cost_id)
			if lines_without_tgv:
				raise UserError('El Gasto Vinculado "%s" tiene Facturas que no tienen Tipo de Gasto Vinculado, DEBE CONFIGURARLO ANTES DE GENERAR EL ASIENTO.'%landed.name)
			self.env.cr.execute(landed.get_sql_saldos())
			res = self.env.cr.dictfetchall()
			data = [p for p in res if p['almacen']]			
			lines = []
			alb = None
			fecha_alb = None
			for l in data:
				if not alb:
					alb = l['albaran']
					fecha_alb = l['fecha_albaran']
				if alb != l['albaran']:
					asiento = self.env['account.move'].create({
						'company_id': landed.company_id.id,
						'journal_id': param.lc_journal_id.id,
						'date': fecha_alb,
						'landed_cost_id':landed.id,
						'glosa': u'POR EL INGRESO A ALMACEN SEGÚN GASTO VINCULADO NRO %s %s'%(landed.name, alb),
						'ref': landed.name, 
						'line_ids':lines})

					asiento._post()
					lines = []
					alb = l['albaran']
					fecha_alb = l['fecha_albaran']
				stock_moveid = self.env['stock.move'].browse(l['stock_moveid'])
				if not stock_moveid.product_id.categ_id.property_stock_valuation_account_id:
					raise UserError(u'La categoría %s no tiene configurada la Cuenta de valoración de stock.'%(stock_moveid.product_id.categ_id.name))
				vals = (0,0,{
				'name': landed.name,
				'debit': l['valor_p'],
				'credit': 0, 
				'account_id': stock_moveid.product_id.categ_id.property_stock_valuation_account_id.id,
				'type_document_id': param.others_document_type.id,
				'product_id': stock_moveid.product_id.id,
				'product_uom_id': stock_moveid.product_uom.id,
				'quantity': l['cantidad'],
				'nro_comp': l['albaran'],
				'company_id': landed.company_id.id,
				})
				lines.append(vals)

				vals = (0,0,{
				'name': landed.name,
				'debit': 0,
				'credit':  l['valor_p'], 
				'account_id': param.lc_account_id.id,
				'type_document_id': param.others_document_type.id,
				'product_id': stock_moveid.product_id.id,
				'product_uom_id': stock_moveid.product_uom.id,
				'quantity': l['cantidad'],
				'nro_comp': l['albaran'],
				'company_id': landed.company_id.id,
				})
				lines.append(vals)

			asiento = self.env['account.move'].create({
				'company_id': landed.company_id.id,
				'journal_id': param.lc_journal_id.id,
				'date': fecha_alb,
				'landed_cost_id':landed.id,
				'glosa': u'POR EL INGRESO A ALMACEN SEGÚN GASTO VINCULADO NRO %s %s'%(landed.name, alb),
				'ref': landed.name, 
				'line_ids':lines})

			asiento._post()

			total_Advalorem = sum(l['Advalorem '] for l in data)
			if total_Advalorem != 0:
				lines = []
				data_Advalorem = [p for p in res if p['Advalorem '] != 0]
				for a in data_Advalorem:
					vals = (0,0,{
					'name': landed.name,
					'debit': a['Advalorem '],
					'credit': 0, 
					'account_id': stock_moveid.product_id.categ_id.property_stock_valuation_account_id.id,
					'type_document_id': param.others_document_type.id,
					'product_id': stock_moveid.product_id.id,
					'product_uom_id': stock_moveid.product_uom.id,
					'quantity': 0,
					'nro_comp': landed.name,
					'company_id': landed.company_id.id,
					})
					lines.append(vals)

					vals = (0,0,{
					'name': landed.name,
					'debit': 0,
					'credit':  a['Advalorem '], 
					'account_id': param.lc_account_id.id,
					'type_document_id': param.others_document_type.id,
					'product_id': stock_moveid.product_id.id,
					'product_uom_id': stock_moveid.product_uom.id,
					'quantity': 0,
					'nro_comp': landed.name,
					'company_id': landed.company_id.id,
					})
					lines.append(vals)

				asiento = self.env['account.move'].create({
					'company_id': landed.company_id.id,
					'journal_id': param.lc_journal_id.id,
					'date': landed.date_kardex.date(),
					'landed_cost_id':landed.id,
					'glosa': u'POR EL INGRESO A ALMACEN SEGÚN GASTO VINCULADO NRO %s ADVALOREM'%(landed.name),
					'ref': landed.name, 
					'line_ids':lines})
				asiento._post()

			for tgv in self.env['landed.cost.it.type'].search([]):
				total_tgv = sum(l[tgv.name] for l in data)
				if total_tgv != 0:
					lines = []
					data_tgv = [p for p in res if p[tgv.name] != 0]
					for a in data_tgv:
						vals = (0,0,{
						'name': landed.name,
						'debit': a[tgv.name],
						'credit': 0, 
						'account_id': stock_moveid.product_id.categ_id.property_stock_valuation_account_id.id,
						'type_document_id': param.others_document_type.id,
						'product_id': stock_moveid.product_id.id,
						'product_uom_id': stock_moveid.product_uom.id,
						'quantity': 0,
						'nro_comp': landed.name,
						'company_id': landed.company_id.id,
						})
						lines.append(vals)

						vals = (0,0,{
						'name': landed.name,
						'debit': 0,
						'credit':  a[tgv.name], 
						'account_id': param.lc_account_id.id,
						'type_document_id': param.others_document_type.id,
						'product_id': stock_moveid.product_id.id,
						'product_uom_id': stock_moveid.product_uom.id,
						'quantity': 0,
						'nro_comp': landed.name,
						'company_id': landed.company_id.id,
						})
						lines.append(vals)

					asiento = self.env['account.move'].create({
						'company_id': landed.company_id.id,
						'journal_id': param.lc_journal_id.id,
						'date': landed.date_kardex.date(),
						'landed_cost_id':landed.id,
						'glosa': u'POR EL INGRESO A ALMACEN SEGÚN GASTO VINCULADO NRO %s %s'%(landed.name,tgv.name),
						'ref': landed.name, 
						'line_ids':lines})
					asiento._post()